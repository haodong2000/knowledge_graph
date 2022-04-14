# -*- utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from sklearn.metrics import brier_score_loss
from xlwt import Workbook
import googletrans
from googletrans import Translator as Google_Translator
from translate import Translator
import time
from requests.packages.urllib3.poolmanager import PoolManager
from requests.adapters import HTTPAdapter
import requests
import ssl


class MyAdapter(HTTPAdapter):
    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = PoolManager(num_pools=connections,
                                       maxsize=maxsize,
                                       block=block,
                                       ssl_version=ssl.PROTOCOL_TLSv1)


class config:
    def __init__(self, filename_node, filename_relation, sheet_names,
                 en_name_col, cn_name_col, type_col, class_col,
                 ori_rel_sh, tar_rel_sh, sheet_overall):
        self.filename_node = filename_node
        self.filename_relation = filename_relation
        self.sheet_names = sheet_names
        self.en_name_col = en_name_col
        self.cn_name_col = cn_name_col
        self.type_col = type_col
        self.class_col = class_col
        self.ori_rel_sh = ori_rel_sh
        self.tar_rel_sh = tar_rel_sh
        self.sheet_overall = sheet_overall


class data_processor:
    def __init__(self, config):
        self.config = config
        self.translator = Translator(from_lang="chinese", to_lang="english")
        self.google = True
        self.google_tran = Google_Translator()
        self.src_language = "zh-cn"
        self.dest_language = "en"

    def _uppercase_no_spaces(self, original):
        en_value = ""
        for c in original:
            if c == ' ':
                en_value += '_'
            elif (c > 'Z' or c < 'A') and (c > 'z' or c < 'a'):
                en_value += '_'
            else:
                en_value += c  # .upper()
        return en_value

    def _translate(self, cn_value, google=False):
        try:
            if google:
                en_values = self.google_tran.translate(cn_value,
                                                       src=self.src_language,
                                                       dest_language=self.dest_language).text
            else:
                en_values = self.translator.translate(cn_value)
            return en_values.capitalize()
        except:
            print("ERROR: failed translating!")
            return ""

    def translate_to_en(self):
        start_time = time.time()
        print("translating start at ->", start_time)
        xlsx = load_workbook(self.config.filename_node)
        for sheet_name in self.config.sheet_names:
            print("translating sheet ->", sheet_name)
            xlsx_sheet = xlsx[sheet_name]
            data = pd.read_excel(self.config.filename_node,
                                 sheet_name=sheet_name)
            col_idx_cn = list(data.columns).index(self.config.cn_name_col) + 1
            col_idx_en = list(data.columns).index(self.config.en_name_col) + 1
            row_num = data.shape[0] + 1
            for i in range(2, row_num + 1):
                if xlsx_sheet.cell(row=i, column=col_idx_en).value != None:
                    continue
                cn_value = xlsx_sheet.cell(row=i, column=col_idx_cn).value
                en_value = self._translate(cn_value, google=self.google)
                while en_value == "":
                    time.sleep(0.5)
                    en_value = self._translate(cn_value, google=self.google)
                en_value = self._uppercase_no_spaces(en_value)
                print(cn_value, "->", en_value)
                xlsx_sheet.cell(row=i, column=col_idx_en).value = en_value
        xlsx.save(config.filename_node)
        print("translating done after ->", time.time() - start_time, "s")

    def intrgrate_all_nodes(self):
        self.translate_to_en()
        start_time = time.time()
        print("intrgrating start at ->", start_time)
        xlsx = load_workbook(self.config.filename_node)
        overall_data = pd.read_excel(
            self.config.filename_node, sheet_name=self.config.sheet_overall)
        overall_sheet = xlsx[self.config.sheet_overall]
        overall_en_col = list(overall_data.columns).index(
            self.config.en_name_col) + 1
        overall_type_col = list(overall_data.columns).index(
            self.config.type_col) + 1
        overall_class_col = list(overall_data.columns).index(
            self.config.class_col) + 1
        overall_index = 2
        for sheet_name in self.config.sheet_names:
            print("translating sheet ->", sheet_name)
            xlsx_sheet = xlsx[sheet_name]
            data = pd.read_excel(self.config.filename_node,
                                 sheet_name=sheet_name)
            col_idx_type = list(data.columns).index(self.config.type_col) + 1
            col_idx_class = list(data.columns).index(self.config.class_col) + 1
            col_idx_en = list(data.columns).index(self.config.en_name_col) + 1
            row_num = data.shape[0] + 1
            for i in range(2, row_num + 1):
                type_value = xlsx_sheet.cell(row=i, column=col_idx_type).value
                class_value = xlsx_sheet.cell(
                    row=i, column=col_idx_class).value
                en_value = xlsx_sheet.cell(row=i, column=col_idx_en).value
                overall_sheet.cell(row=overall_index,
                                   column=overall_type_col).value = type_value
                overall_sheet.cell(row=overall_index,
                                   column=overall_class_col).value = class_value
                overall_sheet.cell(row=overall_index,
                                   column=overall_en_col).value = en_value
                print("current data ->", en_value, type_value, overall_index)
                overall_index += 1
        xlsx.save(config.filename_node)
        print("intrgrating done after ->", time.time() - start_time, "s")

    def _permutations(self, subject, object_):
        match_sub, match_ob = [], []
        for sub in subject:
            for ob in object_:
                match_sub.append(sub)
                match_ob.append(ob)
        return match_sub, match_ob

    def generate_relations(self):
        self.intrgrate_all_nodes()
        start_time = time.time()
        print("generating relations start at ->", start_time)
        xlsx = load_workbook(self.config.filename_relation)
        data = pd.read_excel(self.config.filename_relation,
                             sheet_name=self.config.ori_rel_sh)
        overall_data = pd.read_excel(
            self.config.filename_node, sheet_name=self.config.sheet_overall)
        ori_LABELS = [overall_data[self.config.en_name_col][i]
                      for i in range(0, overall_data.shape[0])]
        ori_TYPES = [overall_data[self.config.type_col][i]
                     for i in range(0, overall_data.shape[0])]
        for label in ori_LABELS:
            if label in ori_TYPES:
                print("ERROR: overlap between labels and types!")
        ori_CLASSES = [overall_data[self.config.class_col][i]
                       for i in range(0, overall_data.shape[0])]
        ori_OVERALL = [[ori_LABELS[i], ori_TYPES[i], ori_CLASSES[i]]
                       for i in range(overall_data.shape[0])]
        OVERALL = []
        [OVERALL.append(item) for item in ori_OVERALL if item not in OVERALL]
        for item in OVERALL:
            print(item)
        index = 2
        for i in range(2, 2 + data.shape[0]):
            sub, rel, ob = xlsx[self.config.ori_rel_sh].cell(row=i, column=1).value, \
                xlsx[self.config.ori_rel_sh].cell(row=i, column=3).value, \
                xlsx[self.config.ori_rel_sh].cell(row=i, column=4).value
            sub_cl, ob_cl = xlsx[self.config.ori_rel_sh].cell(row=i, column=2).value, \
                xlsx[self.config.ori_rel_sh].cell(row=i, column=5).value
            print(sub, sub_cl, rel, ob, ob_cl)
            if sub in ori_TYPES:
                SUB = [OVERALL[i]
                       for i in range(len(OVERALL)) if OVERALL[i][1] == sub and OVERALL[i][2] == sub_cl]
            else:
                SUB = [OVERALL[i]
                       for i in range(len(OVERALL)) if OVERALL[i][0] == sub and OVERALL[i][2] == sub_cl]
                SUB = [SUB[0]]
            if ob in ori_TYPES:
                OB = [OVERALL[i]
                      for i in range(len(OVERALL)) if OVERALL[i][1] == ob and OVERALL[i][2] == ob_cl]
            else:
                OB = [OVERALL[i]
                      for i in range(len(OVERALL)) if OVERALL[i][0] == ob and OVERALL[i][2] == ob_cl]
                OB = [OB[0]]
            match_sub, match_ob = self._permutations(SUB, OB)
            # print(match_sub)
            # print(match_ob)
            for i in range(len(match_sub)):
                for j in range(3):
                    xlsx[self.config.tar_rel_sh].cell(
                        row=index, column=j + 1).value = match_sub[i][j]
                xlsx[self.config.tar_rel_sh].cell(
                        row=index, column=4).value = rel
                for j in range(3):
                    xlsx[self.config.tar_rel_sh].cell(
                        row=index, column=j + 5).value = match_ob[i][j]
                index += 1
        xlsx.save(self.config.filename_relation)
        print("generating relation done after ->",
              time.time() - start_time, "s")


if __name__ == "__main__":
    __s = requests.Session()
    __s.mount('https://', MyAdapter())
    config = config(
        filename_node="./data/insomnia_and_sleep_quality_node.xlsx",
        filename_relation="./data/insomnia_and_sleep_quality_relation.xlsx",
        sheet_names=["MAIN", "CHRONIC_INSOMNIA",
                     "MENSTRUAL_INSOMNIA", "INSOMNIA", "MENOPAUSE_INSOMNIA",
                     "GAD_7", "ISI", "ANALYSIS"],
        en_name_col="EnglishName",
        cn_name_col="ChineseName",
        type_col="Type",
        class_col="Classification",
        ori_rel_sh="RELATION",
        tar_rel_sh="EN_RELATION",
        sheet_overall="OVERALL"
    )
    data_processor = data_processor(config)
    data_processor.generate_relations()
