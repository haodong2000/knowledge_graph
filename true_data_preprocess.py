# -*- utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook
from sklearn.metrics import brier_score_loss
from xlwt import Workbook
import googletrans
from googletrans import Translator as Google_Translator
from translate import Translator
import time
from requests.packages.urllib3.poolmanager import PoolManager
from requests.adapters import HTTPAdapter
import requests
import ssl


class MyAdapter(HTTPAdapter):
    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = PoolManager(num_pools=connections,
                                       maxsize=maxsize,
                                       block=block,
                                       ssl_version=ssl.PROTOCOL_TLSv1)


class config:
    def __init__(self, filename_node, filename_relation, sheet_names, en_name_col, cn_name_col, type_col):
        self.filename_node = filename_node
        self.filename_relation = filename_relation
        self.sheet_names = sheet_names
        self.en_name_col = en_name_col
        self.cn_name_col = cn_name_col
        self.type_col = type_col


class data_processor:
    def __init__(self, config):
        self.config = config
        self.translator = Translator(from_lang="chinese", to_lang="english")
        self.google = True
        self.google_tran = Google_Translator()
        self.src_language = "zh-cn"
        self.dest_language = "en"

    def _uppercase_no_spaces(self, original):
        en_value = ""
        for c in original:
            if c == ' ':
                en_value += '_'
            else:
                en_value += c  # .upper()
        return en_value

    def _translate(self, cn_value, google=False):
        try:
            if google:
                en_values = self.google_tran.translate(cn_value,
                                                       src=self.src_language,
                                                       dest_language=self.dest_language).text
            else:
                en_values = self.translator.translate(cn_value)
            return en_values.capitalize()
        except:
            print("ERROR translating!")
            return ""

    def translate_to_en(self):
        start_time = time.time()
        print("translating start at ->", start_time)
        xlsx = load_workbook(self.config.filename_node)
        for sheet_name in self.config.sheet_names:
            print("translating sheet ->", sheet_name)
            xlsx_sheet = xlsx[sheet_name]
            data = pd.read_excel(self.config.filename_node,
                                 sheet_name=sheet_name)
            col_idx_cn = list(data.columns).index(self.config.cn_name_col) + 1
            col_idx_en = list(data.columns).index(self.config.en_name_col) + 1
            row_num = data.shape[0] + 1
            for i in range(1, row_num + 1):
                if xlsx_sheet.cell(row=i, column=col_idx_en).value != None:
                    continue
                cn_value = xlsx_sheet.cell(row=i, column=col_idx_cn).value
                en_value = self._translate(cn_value, google=self.google)
                while en_value == "":
                    time.sleep(0.5)
                    en_value = self._translate(cn_value, google=self.google)
                en_value = self._uppercase_no_spaces(en_value)
                print(cn_value, "->", en_value)
                xlsx_sheet.cell(row=i, column=col_idx_en).value = en_value
        xlsx.save(config.filename_node)
        print("translating done after ->", time.time() - start_time)


if __name__ == "__main__":
    __s = requests.Session()
    __s.mount('https://', MyAdapter())
    config = config(
        filename_node="./data/insomnia_and_sleep_quality_node.xlsx",
        filename_relation="./data/insomnia_and_sleep_quality_relation.xlsx",
        sheet_names=["MAIN", "CHRONIC_INSOMNIA",
                     "MENSTRUAL_INSOMNIA", "INSOMNIA", "MENOPAUSE_INSOMNIA"],
        en_name_col="EnglishName",
        cn_name_col="ChineseName",
        type_col="Type"
    )
    data_processor = data_processor(config)
    data_processor.translate_to_en()
