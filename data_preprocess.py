# -*- utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
from xlwt import Workbook


def no_space(name):
    no_space = ""
    for c in name:
        if c == ' ':
            no_space += '_'
        else:
            no_space += c
    return no_space

def create_sheet(sheet_name, savename, tran_dict, ori_data, type_dict):
    wb = Workbook()
    data_sheet = wb.add_sheet(sheet_name)
    data_sheet.write(0, 0, "subject")
    data_sheet.write(0, 1, "sub_type")
    data_sheet.write(0, 2, "relation")
    data_sheet.write(0, 3, "object")
    data_sheet.write(0, 4, "obj_type")
    for i in range(len(ori_data)):
        data_sheet.write(i + 1, 0, no_space(tran_dict[ori_data[ori_data.columns[0]][i]]))
        data_sheet.write(i + 1, 1, no_space(type_dict[tran_dict[ori_data[ori_data.columns[0]][i]]]))
        data_sheet.write(i + 1, 2, no_space(ori_data[ori_data.columns[1]][i]))
        data_sheet.write(i + 1, 3, no_space(tran_dict[ori_data[ori_data.columns[2]][i]]))
        data_sheet.write(i + 1, 4, no_space(type_dict[tran_dict[ori_data[ori_data.columns[2]][i]]]))
    wb.save(savename)
    print("sheet", sheet_name, "created")


def read_data(filename, sheet_name="node"):
    print("read data ->", filename, sheet_name)
    return pd.read_excel(filename, sheet_name=sheet_name)


def build_dictionary(data, cn_label, en_label, type_label):
    cn_names, en_names, types = data[cn_label], data[en_label], data[type_label]
    print("length of names ->", len(cn_names), len(en_names))
    tran_dict, type_dict = {}, {}
    for i in range(len(cn_names)):
        tran_dict[cn_names[i]] = en_names[i]
        type_dict[en_names[i]] = types[i]
    return tran_dict, type_dict


if __name__ == "__main__":
    filename = "./data/exemplar_knowledge_graph_v1.xlsx"
    data = read_data(filename)
    cn_label = "ChineseName"
    en_label = "EnglishName"
    type_label = "Type"
    tran_dict, type_dict = build_dictionary(data, cn_label, en_label, type_label)
    savename = "./data/en_relation.xls"
    ori_data = read_data(filename, sheet_name="relation")
    create_sheet("en_relation", savename, tran_dict, ori_data, type_dict)
